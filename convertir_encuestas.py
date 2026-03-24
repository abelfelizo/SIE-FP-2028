#!/usr/bin/env python3
"""
SIE 2028 — Conversor de Plantilla Excel a JSON
================================================
Lee la plantilla sie2028_encuestas_template.xlsx y genera
data/encuestas_sie2028.json listo para cargar al sistema.

Uso:
    python3 convertir_encuestas.py
    python3 convertir_encuestas.py --input mi_archivo.xlsx
    python3 convertir_encuestas.py --input encuestas.xlsx --output data/encuestas.json

El JSON generado tiene el formato exacto que espera MotorEncuestas v10:
    [{
        "fecha":        "2027-03-15",
        "firma":        "Gallup RD",
        "tipo":         "INTENCION_CANDIDATO" | "SIMPATIA_PARTIDARIA",
        "cobertura":    "NACIONAL" | "PROVINCIAL",
        "provincia":    "Distrito Nacional",   (solo si provincial)
        "provincia_id": "01",                  (solo si provincial)
        "n":            1200,
        "calidad":      "A",
        "PRM":          52.0,
        "FP":           34.0,
        "PLD":          9.0,
        "PRD":          2.5,
        "PCR":          1.5,
        "indecisos":    1.0,
        "nota":         "Primera oleada Q1 2027"
    }]
"""

import argparse
import json
import os
import sys
from datetime import datetime

def convert(input_path, output_path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERROR: openpyxl no instalado. Ejecutar: pip install openpyxl")
        sys.exit(1)

    if not os.path.exists(input_path):
        print(f"ERROR: No se encontró el archivo {input_path}")
        sys.exit(1)

    wb = load_workbook(input_path, data_only=True)

    if 'Encuestas' not in wb.sheetnames:
        print("ERROR: El archivo no contiene la hoja 'Encuestas'")
        sys.exit(1)

    ws = wb['Encuestas']

    encuestas = []
    errores   = []
    omitidas  = 0

    # Fila 3 = encabezados, datos desde fila 4
    for row_idx in range(4, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=c).value for c in range(1, 20)]

        # Columnas (0-indexed):
        # 0=ID, 1=Fecha, 2=Firma, 3=Tipo, 4=Cobertura, 5=Provincia, 6=ProvID,
        # 7=N, 8=Calidad, 9=PRM, 10=FP, 11=PLD, 12=PRD, 13=PCR, 14=Indecisos,
        # 15=Suma(auto), 16=Válida(auto), 17=Nota, 18=URL

        firma = row[2]
        if not firma or str(firma).strip() == '':
            omitidas += 1
            continue

        # Validar fecha
        fecha = row[1]
        if fecha is None:
            errores.append(f"Fila {row_idx}: Fecha vacía en encuesta de '{firma}'")
            continue
        if hasattr(fecha, 'strftime'):
            fecha_str = fecha.strftime('%Y-%m-%d')
        else:
            try:
                datetime.strptime(str(fecha), '%Y-%m-%d')
                fecha_str = str(fecha)
            except ValueError:
                errores.append(f"Fila {row_idx}: Fecha inválida '{fecha}' — usar formato YYYY-MM-DD")
                continue

        # Tipo
        tipo_raw = str(row[3] or '').strip().upper().replace(' ', '_')
        if tipo_raw not in ('INTENCION_CANDIDATO', 'SIMPATIA_PARTIDARIA'):
            errores.append(f"Fila {row_idx}: Tipo inválido '{row[3]}' — usar INTENCION_CANDIDATO o SIMPATIA_PARTIDARIA")
            continue
        tipo = tipo_raw.lower()

        # Cobertura
        cobertura_raw = str(row[4] or 'NACIONAL').strip().upper()
        cobertura = cobertura_raw.lower()

        # Calidad
        calidad = str(row[8] or 'B').strip()
        if calidad not in ('A+', 'A', 'B', 'C', 'D'):
            calidad = 'B'

        # Porcentajes — limpiar None y convertir
        def parse_pct(val):
            if val is None or str(val).strip() == '':
                return None
            try:
                f = float(val)
                return round(f, 2) if 0 <= f <= 100 else None
            except (ValueError, TypeError):
                return None

        prm  = parse_pct(row[9])
        fp   = parse_pct(row[10])
        pld  = parse_pct(row[11])
        prd  = parse_pct(row[12])
        pcr  = parse_pct(row[13])
        indc = parse_pct(row[14])

        if prm is None or fp is None:
            errores.append(f"Fila {row_idx}: PRM o FP vacío en encuesta de '{firma}' ({fecha_str})")
            continue

        enc = {
            'fecha':    fecha_str,
            'firma':    str(firma).strip(),
            'tipo':     tipo,
            'cobertura': cobertura,
            'n':        int(row[7]) if row[7] else 600,
            'calidad':  calidad,
            'PRM':      prm,
            'FP':       fp,
        }
        if pld  is not None: enc['PLD']       = pld
        if prd  is not None: enc['PRD']       = prd
        if pcr  is not None: enc['PCR']       = pcr
        if indc is not None: enc['indecisos'] = indc

        # Provincial — requiere provincia_id
        if cobertura == 'provincial':
            prov_nombre = row[5]
            prov_id     = row[6]
            if not prov_id:
                errores.append(f"Fila {row_idx}: Encuesta provincial sin Prov ID para '{firma}' ({fecha_str})")
                continue
            enc['provincia']    = str(prov_nombre or '').strip()
            enc['provincia_id'] = str(prov_id).strip().zfill(2)

        if row[17]:
            enc['nota'] = str(row[17]).strip()

        # Advertencia si suma muy diferente de 100
        suma = sum(v for v in [prm, fp, pld, prd, pcr] if v is not None)
        if abs(suma - 100) > 3:
            print(f"⚠️  Fila {row_idx} '{firma}': suma de partidos = {suma:.1f}% (esperado ~100%)")

        encuestas.append(enc)

    # Ordenar por fecha descendente (más reciente primero)
    encuestas.sort(key=lambda x: x['fecha'], reverse=True)

    # Crear directorio si no existe
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(encuestas, f, ensure_ascii=False, indent=2)

    # Reporte
    print(f"\n{'='*55}")
    print(f"  SIE 2028 — Conversión de Encuestas")
    print(f"{'='*55}")
    print(f"  Archivo fuente:  {input_path}")
    print(f"  Archivo destino: {output_path}")
    print(f"  Encuestas válidas exportadas: {len(encuestas)}")

    candidato = [e for e in encuestas if e['tipo'] == 'intencion_candidato']
    simpatia  = [e for e in encuestas if e['tipo'] == 'simpatia_partidaria']
    prov_enc  = [e for e in encuestas if e['cobertura'] == 'provincial']

    print(f"  └─ Intención candidato (presidencial): {len(candidato)}")
    print(f"  └─ Simpatía partidaria (legislativo):  {len(simpatia)}")
    print(f"  └─ Encuestas provinciales:             {len(prov_enc)}")
    if omitidas:
        print(f"  Filas vacías omitidas: {omitidas}")
    if errores:
        print(f"\n  ⚠️  ERRORES ({len(errores)}):")
        for e in errores:
            print(f"     • {e}")
    print(f"{'='*55}")
    print(f"\n  ✅ Listo. Para cargar al SIE 2028:")
    print(f"     Opción A: El sistema detecta {output_path} automáticamente al arrancar.")
    print(f"     Opción B: Botón 'Cargar JSON' en la vista Encuestas del sistema.\n")

    return len(encuestas)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Convierte la plantilla Excel de encuestas al formato JSON del SIE 2028'
    )
    parser.add_argument('--input',  default='sie2028_encuestas_template.xlsx',
                        help='Ruta al archivo Excel (default: sie2028_encuestas_template.xlsx)')
    parser.add_argument('--output', default='data/encuestas_sie2028.json',
                        help='Ruta al JSON de salida (default: data/encuestas_sie2028.json)')
    args = parser.parse_args()
    convert(args.input, args.output)
